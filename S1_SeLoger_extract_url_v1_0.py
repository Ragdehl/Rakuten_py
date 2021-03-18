# -*- coding: utf-8 -*-
"""
Created on Sat May 12 00:17:02 2018

This code extracts all the links of the existing announces of appartments in Toulouse on the site SeLoger.fr

@author: Edgar
"""
from bs4 import BeautifulSoup
import json
import datetime as dt
import time as tm
import random
from get_proxies import get_proxies
import openpyxl
from hours_difference import hours_difference

Today = dt.datetime.fromtimestamp(tm.time())
today_string = '{0:%Y-%m-%d %H:%M:%S}'.format(dt.datetime.now())
Yesterday = Today - dt.timedelta(days = 1)

with open('SeLoger_url.json', 'r') as read_file:#open the dictionary where all the url's are stored
    url_stored = json.load(read_file)
    read_file.close()

book = openpyxl.load_workbook('S1_SeLoger_extract_url_log.xlsx')
sheet = book.active
last_scrap = sheet.cell(row = sheet.max_row, column = 1).value
book.save('S1_SeLoger_extract_url_log.xlsx')

list_url = ['https://www.seloger.com/list.htm?enterprise=0&natures=1,2,4&places=[{ci:310555}]&projects=2,5&qsversion=1.0&sort=d_dt_crea&types=1,2&LISTING-LISTpg=1']
generic_url = 'https://www.seloger.com/list.htm?enterprise=0&natures=1,2,4&places=%5b%7bci%3a310555%7d%5d&projects=2,5&qsversion=1.0&sort=d_dt_crea&types=1,2&LISTING-LISTpg={}'

for number in range(2, 10):
    list_url.append(generic_url.format(number))#list of urls that contains the announces urls

out = 0
count = 0
new_url = {}

for url in list_url:#go through each page of SeLoger: each one contains arround 35 announces url which have to be scrapped
    print('{0:%Y-%m-%d %H:%M:%S}'.format(dt.datetime.now()) + ' [S1_SeLoger_extract_url]: Scrapping page number ' + str(count))
    while (1):
        page = get_proxies(url)
        soup = BeautifulSoup(page.content, 'html.parser')
        list_announce = soup.find_all('a',class_='c-pa-link link_AB')#each <a> tag contains an url to an announce 
        if list_announce != []:
            break
        else:
            soup_fail = soup
            print('{0:%Y-%m-%d %H:%M:%S}'.format(dt.datetime.now()) + ' [S1_SeLoger_extract_url]: Content was empty, retry with new proxy')           
    for announce in list_announce:#go through each <a> tag on the page in order to scrap all the announcements links
        link = announce['href']
        ident = link[link.find('.htm')-9:link.find('.htm')]#each url contains a unic number that can be used as identifier
        site = 'Old'
        
        if int(link.find('.htm')) < 0:#there are two types of links .htm (old site) and /#?cmp (new site)
            ident = link[link.find('/#?cmp')-9:link.find('/#?cmp')]#each url contains a unic number that can be used as identifier
            site = 'New'
            
        if ident in url_stored:#checks if ident already exists
            if last_scrap == url_stored[ident]['Date']:
                out = 1
                print('{0:%Y-%m-%d %H:%M:%S}'.format(dt.datetime.now()) + ' [S1_SeLoger_extract_url]: Ident ' + ident + ' already existing')
                break#there are no more new announces. Then, stop searching!
        
        #Store new idents
        new_url[ident] = {}
        url_stored[ident] = {}
        url_stored[ident]['Link'] = link
        url_stored[ident]['Status'] = 'New'
        url_stored[ident]['Site'] = site
        url_stored[ident]['Date'] = today_string
        print('{0:%Y-%m-%d %H:%M:%S}'.format(dt.datetime.now()) + ' [S1_SeLoger_extract_url]: New ident ' + ident)
    if out == 1:
        break#there are no more new announces. Then, stop searching!
   
    count = count + 1
    
    wait = random.uniform(9, 15)
    print('{0:%Y-%m-%d %H:%M:%S}'.format(dt.datetime.now()) + ' [S1_SeLoger_extract_url]: Wait for ' + str(round(wait,2)) + ' seconds')
    tm.sleep(wait)#random wait time before continue

diff = round(hours_difference(last_scrap, today_string),2)

print ('{0:%Y-%m-%d %H:%M:%S}'.format(dt.datetime.now()) + ' [S1_SeLoger_extract_url]: ' + str(diff) + ' hours since last scrap')

if len(new_url) > 5:
    book = openpyxl.load_workbook('S1_SeLoger_extract_url_log.xlsx')
    sheet = book.active
    max_row = sheet.max_row + 1
    sheet.cell(column = 1, row = max_row, value = today_string)
    sheet.cell(column = 2, row = max_row, value = len(new_url))
    sheet.cell(column = 3, row = max_row, value = diff)
    sheet.cell(column = 4, row = max_row, value = len(url_stored))
    book.save('S1_SeLoger_extract_url_log.xlsx')
    
    with open('SeLoger_url.json', 'w', encoding='utf8') as write_file:#rewrites file with all new announces
        json.dump(url_stored,write_file)
        write_file.close()
        
    print('{0:%Y-%m-%d %H:%M:%S}'.format(dt.datetime.now()) + ' [S1_SeLoger_extract_url]: There have been ' + str(len(new_url)) + ' new announces since last scrap')
    print('{0:%Y-%m-%d %H:%M:%S}'.format(dt.datetime.now()) + ' [S1_SeLoger_extract_url]: There are a total of ' + str(len(url_stored)) + ' announces')

else:
    print('{0:%Y-%m-%d %H:%M:%S}'.format(dt.datetime.now()) + ' [S1_SeLoger_extract_url]: There are only ' + str(len(new_url)) + ' new announces and have not been registered')
    